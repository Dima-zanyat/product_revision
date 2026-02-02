"""
Views для обработки загрузки Excel файлов.
"""

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import Revision, RevisionProductItem
from products.models import Product
from django.db import transaction

try:
    import openpyxl
except ImportError:
    openpyxl = None


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_excel_products(request):
    """
    Загрузить продукты из Excel файла.
    
    Ожидаемые колонки:
    - Номенклатура (название продукта)
    - Количество (количество в штуках)
    """
    if 'file' not in request.FILES:
        return Response(
            {'error': 'Файл не предоставлен'},
            status=status.HTTP_400_BAD_REQUEST
        )

    file = request.FILES['file']
    revision_id = request.data.get('revision')

    if not revision_id:
        return Response(
            {'error': 'Не указана ревизия'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        revision = Revision.objects.get(id=revision_id)
    except Revision.DoesNotExist:
        return Response(
            {'error': 'Ревизия не найдена'},
            status=status.HTTP_404_NOT_FOUND
        )

    if revision.status != 'draft':
        return Response(
            {'error': 'Можно загружать продукты только в ревизию со статусом "Черновик"'},
            status=status.HTTP_400_BAD_REQUEST
        )

    if not openpyxl:
        return Response(
            {'error': 'Библиотека openpyxl не установлена'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        # Найти заголовки
        headers = {}
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            for col_idx, cell_value in enumerate(row, 1):
                if cell_value:
                    cell_str = str(cell_value).strip().lower()
                    if 'номенклатура' in cell_str or 'наименование' in cell_str:
                        headers['product'] = col_idx
                    elif 'количество' in cell_str:
                        headers['quantity'] = col_idx

        if 'product' not in headers or 'quantity' not in headers:
            return Response(
                {'error': 'Не найдены необходимые колонки: Номенклатура и Количество'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Обработать данные
        count = 0
        errors = []

        with transaction.atomic():
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                product_name = row[headers['product'] - 1] if headers['product'] <= len(row) else None
                quantity = row[headers['quantity'] - 1] if headers['quantity'] <= len(row) else None

                if not product_name or not quantity:
                    continue

                try:
                    product_name = str(product_name).strip()
                    quantity = int(float(quantity))

                    # Найти или создать продукт
                    product, created = Product.objects.get_or_create(
                        title=product_name,
                        defaults={'description': ''}
                    )

                    # Создать или обновить элемент ревизии
                    RevisionProductItem.objects.update_or_create(
                        revision=revision,
                        product=product,
                        defaults={'actual_quantity': quantity}
                    )
                    count += 1
                except Exception as e:
                    errors.append(f'Строка {row_idx}: {str(e)}')

        return Response({
            'success': True,
            'count': count,
            'errors': errors[:10] if errors else []
        })

    except Exception as e:
        return Response(
            {'error': f'Ошибка при обработке файла: {str(e)}'},
            status=status.HTTP_400_BAD_REQUEST
        )
